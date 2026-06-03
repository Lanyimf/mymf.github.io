"""
人工總表 雙向同步腳本
  Excel → DB：偵測到 xlsx 變更後，diff 並更新對應資料庫
  DB → Excel：偵測到 DB 變更後，重新匯出對應 sheet

用法：python scripts/sync_spreadsheet.py
"""
import time, hashlib, sqlite3, shutil
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def log_revision(sheet: str, table: str, action: str, row_key,
                 field=None, old_val=None, new_val=None, source="excel"):
    """寫一筆變更記錄到 revision_log"""
    try:
        conn = sqlite3.connect(str(L3_DB))
        conn.execute("""
            INSERT INTO revision_log (sheet, table_name, action, row_key, field, old_value, new_value, source)
            VALUES (?,?,?,?,?,?,?,?)
        """, (sheet, table, action, str(row_key), field,
              str(old_val)[:500] if old_val is not None else None,
              str(new_val)[:500] if new_val is not None else None,
              source))
        conn.commit()
        conn.close()
    except Exception:
        pass

BASE      = Path(__file__).parent.parent
XLSX_PATH = BASE / "人工總表.xlsx"
L1_DB     = BASE / "src" / "layer1_data"  / "parcels.db"
L3_DB     = BASE / "src" / "layer3_output" / "evaluation.db"
POLL_SEC  = 3

# ── Sheet 定義：(sheet名稱, db路徑, table名稱, 主鍵欄位) ──────────
SHEETS = [
    ("L1_場址資料",   L1_DB, "parcels",              "場址編號"),
    ("L1_C項目許可",  L1_DB, "c_items",              "id"),
    ("L1_用地代碼",   L1_DB, "zone_code_map",         "代碼"),
    ("L2_評估條件",   L3_DB, "evaluation_criteria",   "條件代碼"),
    ("L2_財務參數",   L3_DB, "financial_params",      "用途類型"),
    ("L3_評估結果",   L3_DB, "evaluation_results",    "id"),
]

HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")


# ════════════════════════════════════════════════════════════════
# 工具函式
# ════════════════════════════════════════════════════════════════
def md5(path: Path) -> str:
    try:
        return hashlib.md5(path.read_bytes()).hexdigest()
    except Exception:
        return ""


def read_table(db: Path, table: str) -> pd.DataFrame:
    conn = sqlite3.connect(db)
    df = pd.read_sql(f"SELECT * FROM {table}", conn)
    conn.close()
    return df


def write_table(db: Path, table: str, df: pd.DataFrame, pk: str):
    """把 DataFrame 寫回 DB（INSERT OR REPLACE）"""
    conn = sqlite3.connect(db)
    # 取得 DB 現有欄位（避免 df 含 DB 沒有的欄）
    db_cols = [c[1] for c in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    df = df[[c for c in df.columns if c in db_cols]]
    df = df.dropna(subset=[pk])

    inserted = updated = deleted = 0
    existing_ids = set(
        r[0] for r in conn.execute(f"SELECT {pk} FROM {table}").fetchall()
    )
    new_ids = set(df[pk].astype(str))

    # 刪除不再存在的列
    for rid in existing_ids - new_ids:
        conn.execute(f"DELETE FROM {table} WHERE {pk}=?", (rid,))
        deleted += 1

    # INSERT OR REPLACE
    cols = list(df.columns)
    ph = ",".join(["?"] * len(cols))
    for _, row in df.iterrows():
        vals = [None if pd.isna(v) else v for v in row.values]
        conn.execute(
            f"INSERT OR REPLACE INTO {table} ({','.join(cols)}) VALUES ({ph})",
            vals
        )
        if str(row[pk]) in existing_ids:
            updated += 1
        else:
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, updated, deleted


# ════════════════════════════════════════════════════════════════
# DB → Excel 匯出
# ════════════════════════════════════════════════════════════════
def export_all_to_excel():
    """讀取所有 DB，輸出多 sheet Excel"""
    backup = None
    if XLSX_PATH.exists():
        backup = XLSX_PATH.with_suffix(".bak.xlsx")
        shutil.copy2(XLSX_PATH, backup)

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        for sheet_name, db, table, pk in SHEETS:
            try:
                df = read_table(db, table)
            except Exception as e:
                df = pd.DataFrame({"錯誤": [str(e)]})
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 美化表頭
    wb = load_workbook(XLSX_PATH)
    for sheet_name, _, _, _ in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill = fill
        # 自動欄寬
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
    wb.save(XLSX_PATH)

    if backup:
        backup.unlink(missing_ok=True)
    print(f"[{_ts()}] 匯出完成：{XLSX_PATH.name}")


# ════════════════════════════════════════════════════════════════
# Excel → DB 匯入（diff）
# ════════════════════════════════════════════════════════════════
def import_from_excel():
    """讀取 Excel，對每個 sheet diff 並更新 DB"""
    try:
        xl = pd.ExcelFile(XLSX_PATH)
    except Exception as e:
        print(f"  ❌ 無法讀取 Excel：{e}")
        return

    for sheet_name, db, table, pk in SHEETS:
        if sheet_name not in xl.sheet_names:
            continue
        try:
            df_new = xl.parse(sheet_name)
            ins, upd, dlt = write_table(db, table, df_new, pk)
            if ins or upd or dlt:
                print(f"  [{sheet_name}] +{ins} 新增  ~{upd} 更新  -{dlt} 刪除")
        except Exception as e:
            print(f"  [{sheet_name}] ❌ {e}")


# ════════════════════════════════════════════════════════════════
# 主監聽迴圈
# ════════════════════════════════════════════════════════════════
def _ts():
    return datetime.now().strftime("%H:%M:%S")


def watch():
    print(f"🔄 雙向同步監聽啟動")
    print(f"   Excel: {XLSX_PATH.name}")
    print(f"   DB:    parcels.db + evaluation.db")
    print(f"   每 {POLL_SEC}s 掃描一次，Ctrl+C 停止\n")

    # 初次匯出
    print(f"[{_ts()}] 初始匯出所有 DB → Excel...")
    export_all_to_excel()

    xlsx_hash = md5(XLSX_PATH)
    db_hashes = {str(L1_DB): md5(L1_DB), str(L3_DB): md5(L3_DB)}
    pending_xlsx = None

    while True:
        time.sleep(POLL_SEC)

        # 1. 偵測 Excel 變更 → 更新 DB
        new_xlsx_hash = md5(XLSX_PATH)
        if new_xlsx_hash != xlsx_hash:
            xlsx_hash = new_xlsx_hash
            pending_xlsx = time.time()
            print(f"[{_ts()}] Excel 變更偵測，等待 5s 穩定...")

        if pending_xlsx and (time.time() - pending_xlsx) >= 5:
            pending_xlsx = None
            print(f"[{_ts()}] Excel → DB 同步中...")
            import_from_excel()
            # 更新 DB hash（因為我們剛寫入）
            db_hashes = {str(L1_DB): md5(L1_DB), str(L3_DB): md5(L3_DB)}
            print(f"[{_ts()}] ✅ DB 已更新")

        # 2. 偵測 DB 變更 → 更新 Excel
        for db_path in [str(L1_DB), str(L3_DB)]:
            new_h = md5(Path(db_path))
            if new_h != db_hashes.get(db_path):
                db_hashes[db_path] = new_h
                print(f"[{_ts()}] DB 變更偵測，重新匯出 Excel...")
                export_all_to_excel()
                xlsx_hash = md5(XLSX_PATH)
                break


if __name__ == "__main__":
    watch()
